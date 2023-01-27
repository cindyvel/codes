#name = input("What's your name? \n")
#print("Hi, " + name + ". How's your day? ")
#day = input()
#print("So your day is " + day)
#if day == "bad":
#    print("So your day is " + day + ", whatever happens, don't give up.")
#elif day == "boring":
#    print("Remember, you are more likely to get \n in trouble if you're not doing anything.")
#elif day == "good":
#    print("Your day is " + day + "? That's a relief! Hope everything stays like that!")
#else:
#    print("Well, that's... interesting.")

#print("Hello! What's your name?")
#name = input()
#print("Hi, " + str(name) + "! Welcome to the band name generator.\nTo generate the name of your band, \n"
#"I'm going to need you to answer a few questions.")
#city = input("First, what was the name of the city you grew in? \n")
#print("Ok, so your city is, " + str(city) + ".")
#pet = input("Moving on, what is the name of a pet you like? \n")
#print("Excellent! So your band name would be: \n" + str(city) + str(pet))

#print("Insert your bill: \n")
#bill = float(input())
#print("How many people are going to pay? \n")
#people = float(input())
#print("Tip percentage? \n")
#tip = float(input())
#total_tip = round(bill * (tip / 100), 2)
#bill_per_person = round((bill/people) + (total_tip/people), 2)
#print("Total tip is: " + str(total_tip))
#print("Total amount plus tip per person is: " + str(bill_per_person))

# import spotipy
# from spotipy.oauth2 import SpotifyClientCredentials
#
# sp = spotipy.Spotify(auth_manager=SpotifyClientCredentials(client_id="6609e8697cf443ad8bfbe22ca3dbf81e",
#                                                            client_secret="890788c187954e458703d81593bdfff3"))
#
# results = sp.search(q='weezer', limit=20)
# for idx, track in enumerate(results['tracks']['items']):
#     print(idx, track['name'])

import json
import os
from whoisapi import *
client = Client(api_key='at_NkK9aUmd4ZU1QJ8xwqtlVqAXRf0ra')

from openpyxl import load_workbook
wb = load_workbook("IPs.xlsx")  # Work Book
ws = wb['Servers']  # Work Sheet
column = ws['A']  # Column
column_list = [column[x].value for x in range(len(column))]
print(column_list)
print(len(column_list))

ip_address = column_list

try:

   for ipval in range(len(ip_address)):
        # Get parsed whois record as a model instance.
       whois = client.data(ip_address[ipval])
        # Get particular field of the whois record
        #print(whois.created_date_raw)


        # Get raw API response
       resp_str = client.raw_data(ip_address[ipval])

       params = RequestParameters(ip=1, ip_whois=1)

       whois = client.data(ip_address[ipval], params)

        #print(whois.domain_availability_raw)

        # Also you can modify default values of parameters:
       client.parameters.output_format = 'json'
       jsonFile =  client.raw_data((ip_address[ipval]))


       filename=(ip_address[ipval]) + ".json"
       save_dir = r"C:\Users\c.velarde.moreno\PycharmProjects\pythonclasses\jsonips"
       full_path = os.path.join(save_dir, filename)
       with open(full_path, 'w') as f:
           json_string=json.dumps(jsonFile)
           f.write(json_string)
           f.close()

       # print(client.raw_data(ip_address[ipval]))

        filename = "51.46.189.11.json"
        path = r"C:\Users\c.velarde.moreno\PycharmProjects\pythonclasses\jsonips\51.46.189.11.json"

        with open(path) as filename:
           json_data = json.loads(filename.read())
           print(json_data)
           print(type(json_data))
           rows_json = []
           import  json
           for data in json_data:
               domain_name_json = data['domainName']
               print(domain_name_json)
               registrar_Name = data['registrarName']
               contact_Email = data['contactEmail']
               created_Date = data['registryData.createdDate']
               registrant_country = data['registrant.country']
               #print(domain_name)

except ValueError as err:
    print("This one has an error!")
    #with open('failedips.txt', 'w') as f:
       # f.write(str(ip_address[ipval]))
       # f.write('\n')